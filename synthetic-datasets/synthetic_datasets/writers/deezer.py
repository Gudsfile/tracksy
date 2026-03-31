from datetime import datetime
from pathlib import Path

import openpyxl
from tqdm import tqdm

from ..models.deezer import DeezerStreaming

SHEET_NAME = "10_listeningHistory"
COLUMNS = [
    "Song Title",
    "Artist",
    "ISRC",
    "Album Title",
    "IP Address",
    "Listening Time",
    "Platform Name",
    "Platform Model",
    "Date",
]


class DeezerWriter:
    def __init__(self, output_dir: Path, reference_date: datetime) -> None:
        folder = output_dir / "deezer"
        self.xlsx_path_template: str = str(folder) + "/deezer-data_{timestamp}.xlsx"
        self.reference_date = reference_date

    def write(self, records: list[DeezerStreaming]) -> None:
        timestamp = int(self.reference_date.timestamp())
        xlsx_path = Path(self.xlsx_path_template.format(timestamp=timestamp))
        print(f"Write `xlsx` file: status: `starting`, path: `{xlsx_path.absolute()}`, count_records: `{len(records)}`")
        write_xlsx(xlsx_path, records)
        print(f"Write `xlsx` file: status: `success`, path: `{xlsx_path.absolute()}`, count_records: `{len(records)}`")


def write_xlsx(path: Path, streamings: list[DeezerStreaming]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Use a fixed timestamp for deterministic output (openpyxl embeds creation time otherwise)
    _epoch = datetime(1970, 1, 1)
    wb.properties.created = _epoch
    wb.properties.modified = _epoch

    # Simulate real Deezer export structure: other sheets surround the listening history
    wb.create_sheet("00_userProfile")
    ws = wb.create_sheet(SHEET_NAME)

    ws.append(COLUMNS)

    for streaming in tqdm(streamings, desc=f"📦 Writing {path.name}", unit=" records"):
        ws.append(
            [
                streaming.song_title,
                streaming.artist,
                streaming.isrc,
                streaming.album_title,
                streaming.serialize_ip_address(streaming.ip_address),
                str(streaming.listening_time),
                streaming.platform_name,
                streaming.platform_model,
                streaming.serialize_date(streaming.date),
            ]
        )

    wb.create_sheet("20_searchHistory")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

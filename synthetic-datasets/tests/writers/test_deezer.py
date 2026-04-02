import hashlib
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import openpyxl
from pytest import mark

from synthetic_datasets.writers.deezer import COLUMNS, SHEET_NAME, DeezerWriter, write_xlsx


def test_write_xlsx(tmp_path, deezer_streaming_record):
    # given
    out_file = tmp_path / "test.xlsx"
    # when
    write_xlsx(out_file, [deezer_streaming_record], date=datetime(2026, 2, 8))
    # then
    assert out_file.exists()
    wb = openpyxl.load_workbook(out_file)
    assert wb.sheetnames == ["00_userProfile", SHEET_NAME, "20_searchHistory"]
    ws = wb[SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    assert headers == COLUMNS
    assert ws.max_row == 2  # header + 1 data row


def test_write_xlsx_is_deterministic(tmp_path, deezer_streaming_record):
    # given
    path_1 = tmp_path / "test1.xlsx"
    path_2 = tmp_path / "test2.xlsx"
    reference_date = datetime(2026, 2, 8)
    # when
    write_xlsx(path_1, [deezer_streaming_record], reference_date)
    write_xlsx(path_2, [deezer_streaming_record], reference_date)
    # then
    sha1 = hashlib.sha256(open(path_1, "rb").read()).hexdigest()
    sha2 = hashlib.sha256(open(path_2, "rb").read()).hexdigest()
    assert sha1 == sha2


@mark.parametrize("reference_date, row_number", [(datetime(2026, 2, 8), 10), (datetime(2024, 2, 8), 9)])
def test_write_xlsx_row_count(tmp_path, deezer_streaming_record, reference_date, row_number):
    # given
    records = [deezer_streaming_record] * row_number
    out_file = tmp_path / "test.xlsx"
    # when
    write_xlsx(out_file, records, reference_date)
    # then
    wb = openpyxl.load_workbook(out_file)
    ws = wb[SHEET_NAME]
    assert ws.max_row == row_number + 1  # header + data rows


@patch("synthetic_datasets.writers.deezer.write_xlsx")
def test_deezer_writer_write_calls_helper(mock_write_xlsx, tmp_path, deezer_streaming_record):
    # given
    reference_date = datetime(2026, 2, 8)
    writer = DeezerWriter(output_dir=tmp_path, reference_date=reference_date)
    records = [deezer_streaming_record]
    # when
    writer.write(records)
    # then
    assert mock_write_xlsx.called
    expected_timestamp = int(reference_date.timestamp())
    expected_path = Path(tmp_path) / "deezer" / f"deezer-data_{expected_timestamp}.xlsx"
    assert mock_write_xlsx.call_args.args == (expected_path, records, reference_date)
